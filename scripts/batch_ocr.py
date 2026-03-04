import sys
import os
import argparse
import json
import base64
import re
import glob
from typing import Dict, Any, List
from openpyxl import load_workbook

# Add parent directory to path to import skill_base
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from skill_base import agent

EXCEL_PATH = 'results/AI-OCR專案_辨識結果記錄表.xlsx'
SHEET_NAME = '辨識明細表'
IMAGES_ROOT = '/home/ubuntu/Documents/2026-0120_新安東京 - 表單 AI OCR'
REFERENCES_DIR = '/home/ubuntu/ocr_test/skills/form_ocr_skill/references'

def load_field_mappings() -> Dict[str, List[str]]:
    """
    Parses markdown files in references directory to build a map of 
    JSON Key -> List of Chinese Labels.
    """
    mapping = {}
    
    # Regex to find table rows: | `key` | label, label2 | description |
    # Matches: | `key` | labels | ...
    table_row_re = re.compile(r'\|\s*`([^`]+)`\s*\|\s*([^|]+)\s*\|')
    
    for md_file in glob.glob(os.path.join(REFERENCES_DIR, "*.md")):
        with open(md_file, 'r', encoding='utf-8') as f:
            for line in f:
                match = table_row_re.search(line)
                if match:
                    key = match.group(1).strip()
                    labels_str = match.group(2).strip()
                    # Split by comma or Chinese comma
                    labels = [l.strip() for l in re.split(r'[,，]', labels_str) if l.strip()]
                    
                    if key not in mapping:
                        mapping[key] = []
                    # Append unique labels
                    for l in labels:
                        if l not in mapping[key]:
                            mapping[key].append(l)
    return mapping

FIELD_MAPPING = load_field_mappings()

def find_image(file_id: str) -> str:
    """
    Finds the image file for the given ID.
    Logic:
    1. Check if ID has 'B' (Bounding Box), if so, look for 'N' (No BBox) version.
    2. Search recursively in IMAGES_ROOT.
    3. Fallback to original ID if 'N' version not found.
    """
    
    # Logic to switch B to N
    parts = file_id.split('-')
    target_ids = [file_id]
    
    if len(parts) >= 3 and parts[2] == 'B':
        parts[2] = 'N'
        n_id = "-".join(parts)
        # Prioritize N version
        if n_id != file_id:
             target_ids.insert(0, n_id)
             print(f"Will look for switched ID {n_id} first, then {file_id}")

    # Search pattern: recursively find file matching target_id.*
    for target_id in target_ids:
        # We walk once per target_id to be safe, though inconsistent efficiency.
        # Given small number of files, this is acceptable.
        for root, dirs, files in os.walk(IMAGES_ROOT):
            for file in files:
                # Check for exact name match (ignoring extension)
                if file.startswith(target_id) and os.path.splitext(file)[0] == target_id:
                     found_path = os.path.join(root, file)
                     print(f"Found image: {found_path}")
                     return found_path
                     
    return None

def encode_image(image_path):
    """Encodes an image to base64."""
    try:
        with open(image_path, "rb") as image_file:
            return base64.b64encode(image_file.read()).decode('utf-8')
    except Exception as e:
        print(f"Error reading image: {e}")
        return None

def run_ocr(image_path: str) -> Dict[str, Any]:
    """Runs OCR on the given image using the agent."""
    print(f"Processing image: {image_path}")
    
    query = "請幫我提取這張單據的資料，請自行判斷是哪一類單據，然後依照其種類去提取相對應的欄位資料。Output MUST be valid JSON."
    
    base64_image = encode_image(image_path)
    if not base64_image:
        return {}

    message_content = [
        {"type": "text", "text": query},
        {"type": "image_url", "image_url": {"url": f"data:image/jpeg;base64,{base64_image}"}}
    ]

    try:
        result = agent.invoke(
            {"messages": [{"role": "user", "content": message_content}]},
            {"configurable": {"thread_id": "batch_process"}}
        )
        
        last_message = result["messages"][-1].content
        try:
            start = last_message.find('{')
            end = last_message.rfind('}') + 1
            if start != -1 and end != -1:
                json_str = last_message[start:end]
                data = json.loads(json_str)
                return data
            else:
                print("No JSON found in response")
                return {}
        except json.JSONDecodeError:
            print("Failed to decode JSON from response")
            print(last_message)
            return {}
            
    except Exception as e:
        print(f"Error running agent: {e}")
        return {}

def update_excel(excel_path: str, file_id: str, ocr_data: Dict[str, Any]):
    """Updates the Excel file with OCR data using openpyxl."""
    if not ocr_data or 'data' not in ocr_data:
        print("No valid OCR data to update.")
        return

    print(f"Updating Excel for File ID: {file_id}")
    
    try:
        wb = load_workbook(excel_path)
        if SHEET_NAME not in wb.sheetnames:
            print(f"Sheet {SHEET_NAME} not found!")
            return
            
        ws = wb[SHEET_NAME]
        
        # Identify Header Row and Columns (Row 3)
        header_row = 3
        col_map = {}
        
        for cell in ws[header_row]:
            if cell.value:
                col_name = str(cell.value).strip()
                col_map[col_name] = cell.column
        
        required_cols = ['檔案編號', '欄位名稱', '辨識結果']
        for col in required_cols:
            if col not in col_map:
                print(f"Missing column: '{col}' in header row {header_row}. Found: {list(col_map.keys())}")
                return

        file_id_col = col_map['檔案編號']
        field_name_col = col_map['欄位名稱']
        result_col = col_map['辨識結果']
        
        # Find existing rows for this file
        existing_rows = {} # Map '欄位名稱' -> row_index
        
        # Context data to copy
        context_data = {} 
        context_cols = ['專案名稱', 'AI-OCR專案', '辨識作業期間']
        
        for row in range(header_row + 1, ws.max_row + 1):
            fid = ws.cell(row=row, column=file_id_col).value
            if fid == file_id:
                fname = ws.cell(row=row, column=field_name_col).value
                if fname:
                    existing_rows[str(fname).strip()] = row
                
                if not context_data:
                    for cname in context_cols:
                        if cname in col_map:
                             val = ws.cell(row=row, column=col_map[cname]).value
                             context_data[cname] = val

        extracted_data = ocr_data['data']
        
        for key, value in extracted_data.items():
            val_str = str(value) if value is not None else ""
            
            # Find the Chinese field name
            # 1. Check if key matches an existing field name directly? (Unlikely for English key)
            # 2. Use FIELD_MAPPING to find candidates
            candidates = FIELD_MAPPING.get(key, [key])
            
            target_field_name = None
            
            # Try to match with existing rows first
            for candidate in candidates:
                if candidate in existing_rows:
                    target_field_name = candidate
                    break
            
            # If no existing row match, pick the first candidate as the name for the new row
            if not target_field_name:
                target_field_name = candidates[0]

            if target_field_name in existing_rows:
                # Update existing
                row_idx = existing_rows[target_field_name]
                ws.cell(row=row_idx, column=result_col).value = val_str
                print(f"Updated existing {target_field_name} at row {row_idx}: {val_str}")
            else:
                # Append new row
                print(f"Appending new field: {target_field_name} = {val_str}")
                new_row_idx = ws.max_row + 1
                
                ws.cell(row=new_row_idx, column=file_id_col).value = file_id
                ws.cell(row=new_row_idx, column=field_name_col).value = target_field_name
                ws.cell(row=new_row_idx, column=result_col).value = val_str
                
                for cname, cval in context_data.items():
                    if cname in col_map:
                        ws.cell(row=new_row_idx, column=col_map[cname]).value = cval
                        
        wb.save(excel_path)
        print("Excel updated successfully.")

    except Exception as e:
        print(f"Error updating Excel: {e}")
        import traceback
        traceback.print_exc()

import time

def main():
    parser = argparse.ArgumentParser(description="Batch OCR to Excel")
    parser.add_argument("input", help="File ID (e.g., A01-P-B-260116-001) or Path to image")
    args = parser.parse_args()
    
    file_id = ""
    image_path = ""
    
    # Determine if input is path or ID
    if os.path.exists(args.input) or '/' in args.input:
        image_path = args.input
        file_id = os.path.splitext(os.path.basename(image_path))[0]
    else:
        file_id = args.input
        image_path = find_image(file_id)
        
        if not image_path:
            print(f"Could not find image for ID: {file_id}")
            return

    print(f"Target File ID: {file_id}")
    print(f"Using Image: {image_path}")
    
    start_time = time.time()
    ocr_result = run_ocr(image_path)
    end_time = time.time()
    
    duration = end_time - start_time
    print(f"\n--- Performance Metrics ---")
    print(f"Execution Time: {duration:.2f} seconds")
    print(f"Throughput: {1/duration:.4f} calls/sec ({60/duration:.2f} calls/min)")
    print(f"---------------------------")

    print("OCR JSON Result:", json.dumps(ocr_result, indent=2, ensure_ascii=False))
    
    # Save text output
    output_dir = os.path.join(os.path.dirname(EXCEL_PATH), 'ocr_outputs')
    os.makedirs(output_dir, exist_ok=True)
    output_file = os.path.join(output_dir, f"{file_id}.txt")
    with open(output_file, "w", encoding="utf-8") as f:
        json.dump(ocr_result, f, indent=2, ensure_ascii=False)
    print(f"OCR result saved to: {output_file}")
    
    update_excel(EXCEL_PATH, file_id, ocr_result)

if __name__ == "__main__":
    main()
